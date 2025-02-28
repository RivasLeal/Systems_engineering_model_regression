#/usr/bin/env python3
####################################################################################
#   Copyright (c) 2025
#
#   Team: System of a Down
#
#   Authors: Freddie Fraticelli, John Midkiff, Zack Popilek,
#            Ashley Porter, and Gerardo Rivas-Leal
#   
#   
#   This project is open source and can be utilized and referenced
#   as long as the team and all authors are credited.
#
#
#
####################################################################################

from sys import argv
from numpy import log
from Data_Fitting import Model_Fitting
from json import load
from Model_Types.COCOMO import COCOMO_ENUM, COCOMO
from Model_Types.SLIM import SLIM
from random import randint, uniform
from math import ceil
from Generation import create_excel
import numpy as np
from sklearn.linear_model import LinearRegression, Ridge
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import random
import matplotlib.pyplot as plt
import os

# Define the shared variation range
variation_range = 0.10

def main(args):
    # First Argument is the Data JSON File
    with open(r'{}'.format(args[0]), 'r') as data_file:
        data = load(data_file)

    # Sections Found in the JSON
    project_string = "Projects"
    sloc_string = "SLOC"
    dev_time_string = "Development Time"
    effort_string = "Effort"
    labor_hours_string = "Labor Hours in a Month"
    env_mode_string = "Environment Mode"
    gen_var_string = "Selective Generative Variance"
    gener_iter = "Generation Iterations"
    min_sloc = "Min SLOC Generation"
    max_sloc = "MAX SLOC Generation"

    # Storage of the data
    data_array = []
    data_array_temp = []

    all_effort = []
    all_dev_time = []
    all_c = []
    all_C = []
    slim_proj_dict = dict()

    # Sum of all projects
    total_src_code = 0
    total_dev_time = 0
    total_effort = 0

    # Environment Vars found in the JSON
    labor_hours_to_months = data[labor_hours_string]
    env_mode = data[env_mode_string]
    
    gen_var_float = 0.0
    try:
        gen_var_float = float(data[gen_var_string])
    except:
        gen_var_float = 0.15

    # Create a COCOMO ENUM of UNDEFINED
    cocomo_enum = COCOMO_ENUM.UNDEFINED
    cocomo_enum = cocomo_enum.get_model_type_from_string(env_mode)
    cocomo_proj_dict = dict()

    # Iterate through all Projects
    for proj in data[project_string]:
        # Get Values from the json
        sloc = data[project_string][proj][sloc_string]
        dev_time = data[project_string][proj][dev_time_string]
        eff = data[project_string][proj][effort_string]
        slim_proj_dict[proj] = SLIM(env_mode)

        # Convert SLOC -> KSLOC
        cocomo_proj_dict[proj] = COCOMO((sloc / (10**3)), cocomo_enum)

        data_array.append((sloc, dev_time))
        data_array_temp.append((eff, dev_time))
        all_effort.append(eff)
        all_dev_time.append(dev_time)

        slim_proj_dict[proj].S = sloc
        slim_proj_dict[proj].K = eff
        slim_proj_dict[proj].t_d = dev_time
        slim_proj_dict[proj].C = slim_proj_dict[proj].solve_for_constant()

        all_C.append(slim_proj_dict[proj].C)
        all_c.append((sloc / slim_proj_dict[proj].C))

        total_src_code += sloc
        total_dev_time += dev_time
        total_effort += eff

    temp = SLIM(env_mode)
    
    iter = 0
    iter_limit = data[gener_iter]

    # Make sure it cant be non zero
    if (iter_limit < 1):
        iter_limit = 100

    best_p = temp.func.p
    best_q = temp.func.q
    best_c = (sum(all_C) / len(all_C))

    try:
        while (iter < iter_limit):
            all_effort, all_dev_time, all_c, data_array_temp, data_array, \
                total_src_code, total_dev_time, total_effort = get_new_total_values(slim_proj_dict)

            p, q, c = run_analysis(env_mode, all_effort, all_dev_time, all_c, data_array_temp, data_array, 
                total_src_code, total_dev_time, total_effort, slim_proj_dict)
            
            if(c > best_c):
                best_q = q
                best_p = p 
                best_c = c

            iter += 1
    except:
        print("Couldn't continue calculating, using last known P and Q values")

    for proj in slim_proj_dict.values():
        proj.C = best_c
        proj.func.p = best_p
        proj.func.q = best_q
        proj.K = round(proj.solve_for_K(), 4)

    if (len(args) > 1):
        try:
            num_entries = int(args[1])
            print("Generating: {} data points".format(args[1]))
            generate_new_data(data[min_sloc], data[max_sloc], (sum(all_C) / len(all_C)), temp.func.p, temp.func.q, slim_proj_dict, num_entries, total_src_code, total_dev_time, gen_var_float)
        except:
            print("Not a valid int")

    # COCOMO Analysis
    num_entries = int(args[1]) if len(args) > 1 else 0
    if num_entries > 0:
        new_cocomo_data = generate_cocomo_data(data[min_sloc], data[max_sloc], num_entries, total_effort, total_dev_time)
        tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_variance, tuned_time_variance, tuned_effort_prediction, tuned_time_prediction = run_cocomo_analysis(new_cocomo_data, cocomo_enum)
    else:
        tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_variance, tuned_time_variance, tuned_effort_prediction, tuned_time_prediction = run_cocomo_analysis(data, cocomo_enum)
    print("COCOMO Analysis:")
    print(f"  Tuned Effort Variance: {tuned_effort_variance}")
    print(f"  Tuned Time Variance: {tuned_time_variance}")
    print(f"  Tuned COCOMO Coefficients: a={tuned_a}, b={tuned_b}, c={tuned_c}, d={tuned_d}")

    if len(args) > 2:
        createEx = create_excel.CreateSlimExcel(slim_proj_dict, args[2])
        file_loc = args[2] + "UpdatedHistoricalData.xlsx"

        # Prepare data for COCOMO Excel output
        if num_entries > 0:
            cocomo_data_for_excel = new_cocomo_data
        else:
            cocomo_data_for_excel = data

        # Write COCOMO data to the same file
        write_cocomo_to_excel(file_loc, tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_variance, tuned_time_variance, tuned_effort_prediction, tuned_time_prediction, cocomo_data_for_excel)

        # Set the excel path.
        excel_file_path = args[2] + "UpdatedHistoricalData.xlsx"  

        # SLIM Plots
        plot_slim_data(slim_proj_dict, best_c, best_p, best_q, excel_file_path)

        # COCOMO Plots
        if num_entries > 0:
            plot_cocomo_data(new_cocomo_data, tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_prediction, tuned_time_prediction, excel_file_path)
        else:
            plot_cocomo_data(data, tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_prediction, tuned_time_prediction, excel_file_path)
    print("SLIM Analysis:")
    print("New P: {}, New Q: {}".format(best_p, best_q))
 
    # if(len(args) > 2):
    #     createEx = create_excel.CreateSlimExcel(slim_proj_dict, args[2])

    # print("New P: {}, New Q: {}".format(best_p, best_q))

def plot_slim_data(slim_proj_dict, best_c, best_p, best_q, excel_file_path):
    """
    Plots SLIM data and saves plots to an Excel worksheet named 'SLIM Plots'.

    Args:
        slim_proj_dict (dict): Dictionary containing SLIM project data.
        best_c (float): Best calculated technology constant (C).
        best_p (float): Best calculated P value (Gaffney P).
        best_q (float): Best calculated Q value (Gaffney Q).
        excel_file_path (str): Path to the Excel file where plots will be saved.
    """

    slocs = [proj.S for proj in slim_proj_dict.values()]
    efforts = [proj.K for proj in slim_proj_dict.values()]
    dev_times = [proj.t_d for proj in slim_proj_dict.values()]
    ks_p = [proj.K ** best_p for proj in slim_proj_dict.values()]
    tds_q = [proj.t_d ** best_q for proj in slim_proj_dict.values()]

    # Create lists of constant values for P, Q, and C
    ps = [best_p] * len(slocs)
    qs = [best_q] * len(slocs)
    cs = [best_c] * len(slocs)

    plot_data = [
        (slocs, efforts, "Effort vs SLOC", "SLOC", "Effort (Labor Yrs)", "slim_effort_sloc.png"),
        (slocs, dev_times, "Dev Time vs SLOC", "SLOC", "Dev Time (Yrds) t_d", "slim_devtime_sloc.png"),
        (slocs, ks_p, "K^p vs SLOC", "SLOC", "K^p", "slim_kp_sloc.png"),
        (slocs, tds_q, "t_d^q vs SLOC", "SLOC", "t_d^q", "slim_tdq_sloc.png"),
        (slocs, ps, "P vs SLOC", "SLOC", "P", "slim_p_sloc.png"),
        (slocs, qs, "Q vs SLOC", "SLOC", "Q", "slim_q_sloc.png"),
        (slocs, cs, "C vs SLOC", "SLOC", "C", "slim_c_sloc.png"),
    ]

    try:
        workbook = load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    if "SLIM Plots" in workbook.sheetnames:
        del workbook["SLIM Plots"]
    worksheet = workbook.create_sheet("SLIM Plots")

    temp_image_files = []

    for i, (x, y, title, xlabel, ylabel, filename) in enumerate(plot_data):
        plt.figure(figsize=(8, 6))
        plt.scatter(x, y, label="Data Points")

        # Calculate regression line
        coefficients = np.polyfit(x, y, 1)
        poly = np.poly1d(coefficients)
        plt.plot(x, poly(x), color='red', label="Best Fit Line")

        plt.title(title)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.legend()
        plt.grid(True)
        plt.savefig(filename)
        plt.close()

        img = Image(filename)
        worksheet.add_image(img, f"A{i * 30 + 1}")

        temp_image_files.append(filename)

    workbook.save(excel_file_path)

    for filename in temp_image_files:
        os.remove(filename)

def plot_cocomo_data(data, tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_prediction, tuned_time_prediction, excel_file_path):
    """
    Plots COCOMO data and saves plots to an Excel worksheet named 'COCOMO Plots'.

    Args:
        data (dict): Dictionary containing COCOMO project data.
        tuned_a (float): Tuned COCOMO coefficient 'a'.
        tuned_b (float): Tuned COCOMO coefficient 'b'.
        tuned_c (float): Tuned COCOMO coefficient 'c'.
        tuned_d (float): Tuned COCOMO coefficient 'd'.
        tuned_effort_prediction (numpy.ndarray): Predicted effort values.
        tuned_time_prediction (numpy.ndarray): Predicted time values.
        excel_file_path (str): Path to the Excel file where plots will be saved.
    """

    klocs, efforts, dev_times = load_cocomo_data(data)
    slocs = klocs * 1000

    plot_data = [
    (slocs, efforts, "SLOC vs Effort (Labor Months)", "SLOC", "Effort (Labor Months)", "cocomo_sloc_effort.png"),
    (slocs, dev_times, "SLOC vs Dev Time (Months) t_d", "SLOC", "Dev Time (YrMonthss) t_d", "cocomo_sloc_devtime.png"),
    (slocs, tuned_effort_prediction, "SLOC vs Predicted Effort", "SLOC", "Predicted Effort", "cocomo_sloc_predicted_effort.png"),
    (slocs, tuned_time_prediction, "SLOC vs Predicted Time", "SLOC", "Predicted Time", "cocomo_sloc_predicted_time.png"),
    (slocs, [tuned_a] * len(slocs), "SLOC vs COCOMO(a)", "SLOC", "COCOMO(a)", "cocomo_sloc_a.png"),
    (slocs, [tuned_b] * len(slocs), "SLOC vs COCOMO(b)", "SLOC", "COCOMO(b)", "cocomo_sloc_b.png"),
    (slocs, [tuned_c] * len(slocs), "SLOC vs COCOMO(c)", "SLOC", "COCOMO(c)", "cocomo_sloc_c.png"),
]

    try:
        workbook = load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    if "COCOMO Plots" in workbook.sheetnames:
        del workbook["COCOMO Plots"]
    worksheet = workbook.create_sheet("COCOMO Plots")

    temp_image_files = []

    # Calculate regression lines OUTSIDE the loop
    coefficients_effort = np.polyfit(slocs, tuned_effort_prediction, 1)
    poly_effort = np.poly1d(coefficients_effort)

    coefficients_time = np.polyfit(slocs, tuned_time_prediction, 1)
    poly_time = np.poly1d(coefficients_time)

    for i, (x, y, title, xlabel, ylabel, filename) in enumerate(plot_data):
        plt.figure(figsize=(8, 6))
        plt.scatter(x, y, label="Data Points", s=20)

        if title == "SLOC vs Predicted Effort":
            plt.plot(x, poly_effort(x), color='red', label="Best Fit Line", linewidth=1)
        elif title == "SLOC vs Predicted Time":
            plt.plot(x, poly_time(x), color='red', label="Best Fit Line", linewidth=1)
        else:
            coefficients = np.polyfit(x, y, 1)
            poly = np.poly1d(coefficients)
            plt.plot(x, poly(x), color='red', label="Best Fit Line", linewidth=1)

        plt.title(title)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.legend()
        plt.grid(True)

        if title == "SLOC vs Dev Time (Yrds) t_d":
            plt.ylim(0, 10)

        plt.savefig(filename)
        plt.close()

        img = Image(filename)
        worksheet.add_image(img, f"A{i * 30 + 1}")

        temp_image_files.append(filename)

    workbook.save(excel_file_path)

    for filename in temp_image_files:
        os.remove(filename)

def generate_new_data(min_sloc, max_sloc, best_c, best_p, best_q, slim_proj_dict, num_entries, total_src_code, total_dev_time, gen_var_float):
    
    try:
        lowerLimit = int(min_sloc)
    except:
        lowerLimit = 45000
        print("Lower Limit couldn't be read in, using 45000")
    try:
        upperLimit = int(max_sloc)
    except:
        upperLimit = 900000
        print("Upper Limit couldn't be read in, using 900000")

    ratio = total_dev_time / total_src_code
    itr = 1

    while num_entries > 0 and ratio != 0:
        project = "project{}".format(itr)

        slim_proj_dict[project] = SLIM()
        slim_proj_dict[project].S = int(ceil(randint(lowerLimit, upperLimit) / 100.0)) * 100

        estimated_development_time = (ratio* slim_proj_dict[project].S)
        variance = 0
        
        while variance == 0:
           variance = 1 + uniform((-1 *gen_var_float),gen_var_float)

        slim_proj_dict[project].t_d = round(estimated_development_time, 2) 

        slim_proj_dict[project].func.p =  best_p
        slim_proj_dict[project].func.q =  best_q

        slim_proj_dict[project].C =  best_c

        slim_proj_dict[project].K = round(slim_proj_dict[project].solve_for_K() * variance,4)

        itr += 1
        num_entries -=1

def generate_cocomo_data(min_sloc, max_sloc, num_entries, total_effort, total_dev_time):
    """
    Generates random data for COCOMO calculations with updated logic.

    Args:
        min_sloc (int): Minimum Source Lines of Code (SLOC).
        max_sloc (int): Maximum Source Lines of Code (SLOC).
        num_entries (int): Number of data entries to generate.
        total_effort (float): Approximate total effort for the generated projects.
        total_dev_time (float): Approximate total development time for the generated projects.

    Returns:
        dict: A dictionary containing generated project data.
    """
    try:
        lowerLimit = int(min_sloc)
    except ValueError:
        lowerLimit = 45000
        print("Lower Limit couldn't be read in, using 45000")

    try:
        upperLimit = int(max_sloc)
    except ValueError:
        upperLimit = 900000
        print("Upper Limit couldn't be read in, using 900000")

    if num_entries <= 0:
        return {"Projects": {}}

    cocomo_data = {"Projects": {}}
    for i in range(1, num_entries + 1):
        project_name = f"project{i}"
        sloc = int(np.ceil(random.randint(lowerLimit, upperLimit) / 100.0)) * 100

        # Effort depends on SLOC with almost NO randomness
        base_effort = sloc * 0.000016 # Fixed value, no randomness
        effort_variation = random.uniform(-0.005, 0.005) # Tiny variation
        effort = base_effort * (1 + effort_variation)

        # Dev Time depends on SLOC with almost NO randomness
        base_time = sloc * 0.000008 # Fixed value, no randomness
        time_variation = random.uniform(-0.005, 0.005) # Tiny variation
        development_time = base_time * (1 + time_variation)

        effort = max(0.1, effort)
        development_time = max(0.1, development_time)

        cocomo_data["Projects"][project_name] = {
            "SLOC": sloc,
            "Effort": effort,
            "Development Time": development_time,
        }

    return cocomo_data
def get_new_total_values(slim_proj_dict):
    '''
    @param slim_proj_dict   - dictionary of SLIM models

    @return all_effort      - Lists of all projects Effort
    @return all_dev_time    - Lists of all project development time
    @return all_c           - Lists of all projects technology constant
    @return data_array_temp - 2D list containing Effort and development time for each project
    @return data_array      - 2D list Source Lines of Code and development time for each project
    @return total_src_code  - Total number of source code throughout all projects
    @return total_dev_time  - Total number of development time throughout all projects
    @return total_effort    - Total number of effort throughout all projects
    '''
    # Storage of the data
    data_array = []
    data_array_temp = []

    all_effort = []
    all_dev_time = []
    all_c = []

    # Sum of all projects
    total_src_code  = 0
    total_dev_time  = 0
    total_effort    = 0 

    for proj in slim_proj_dict.values():

        data_array.append((proj.S,proj.t_d))
        data_array_temp.append((proj.K, proj.t_d))
        all_effort.append(proj.K)
        all_dev_time.append(proj.t_d)

        total_src_code += proj.S
        total_dev_time += proj.t_d
        total_effort   += proj.K

        all_c.append(proj.S / proj.C)

    return all_effort, all_dev_time, all_c, data_array_temp, data_array, \
            total_src_code, total_dev_time, total_effort

def run_analysis(env_mode, all_effort, all_dev_time, all_c, data_array_temp, data_array, 
              total_src_code, total_dev_time, total_effort, slim_proj_dict):
    '''
    @param env_mode        - Whether we are running Gafney or Putnam
    @param slim_proj_dict  - dictionary of SLIM models
    @param all_effort      - Lists of all projects Effort
    @param all_dev_time    - Lists of all project development time
    @param all_c           - Lists of all projects technology constant
    @param data_array_temp - 2D list containing Effort and development time for each project
    @param data_array      - 2D list Source Lines of Code and development time for each project
    @param total_src_code  - Total number of source code throughout all projects
    @param total_dev_time  - Total number of development time throughout all projects
    @param total_effort    - Total number of effort throughout all projects

    @return newAvgP - New Calculated P value
    @return newAvgQ - New Calculated Q value
    @return tempC   - New Calculated Technology Constant value
    '''
    temp_SLIM = SLIM(env_mode)

    cur_pl = Model_Fitting.Model_Fitting()

    # Fit the Data into a linear model
    a, b = cur_pl.fit_exponential_equation(all_effort, all_dev_time, all_c)

    c = []

    # Get the (S/C)^2 values and save off the values
    for sloc_, dev_ in data_array_temp:
        cal_c = cur_pl.get_C( a, b, sloc_, dev_)
        c.append(cal_c)

    # The values obtained are currently (S/C)^2, so need to convert to S/C
    sqrt_c_vals = []
    for _ in c:
        sqrt_c_vals.append(_ ** 0.5)

    # Get all the C values
    new_C_values = []
    for c , a in zip(sqrt_c_vals,data_array):
       new_C_values.append(a[0] / c)

    # Store the New P and Q values
    newPs = []
    newQs = []

    # Find all the P and Q values from the data
    for newC, arTemp, daTemp in zip(new_C_values, data_array_temp, data_array):
        ln_K = log(arTemp[0])
        ln_td = log(arTemp[1])
        ln_SLOC_newC = log(daTemp[0]/newC)

        temp11 = ln_SLOC_newC - (temp_SLIM.func.p * ln_K)
        temp12 = ln_SLOC_newC - (temp_SLIM.func.q * ln_td)
        
        newPs.append(temp11 / ln_td )
        newQs.append(temp12 / ln_K )

    # Get the average P and Q values for all projects
    newAvgP = round(sum(newPs) / len(newPs), 4)
    newAvgQ = round(sum(newQs) / len(newQs), 4)

    # Get the Average values for all projects to get an overall C value
    SLOC_avg     = total_src_code / len(slim_proj_dict)
    dev_time_avg = total_dev_time / len(slim_proj_dict)
    effort_avg   = total_effort   / len(slim_proj_dict)

    # Plug in Values into SLIM as if it where a solo project
    temp_SLIM.S = SLOC_avg
    temp_SLIM.K = effort_avg
    temp_SLIM.t_d = dev_time_avg
    temp_SLIM.func.p = newAvgP
    temp_SLIM.func.q = newAvgQ

    # Solve for the company C value
    tempC = round(temp_SLIM.solve_for_constant(), 4)

    for proj in slim_proj_dict.values():

        # Insert the New P, Q, and C values into all project and revaluate the effort
        proj.func.p = newAvgP
        proj.func.q = newAvgQ

    return newAvgP, newAvgQ, tempC

def load_cocomo_data(data):
    """Loads COCOMO data and converts SLOC to KLOC."""
    kloc_values = []
    effort_values = []
    time_values = []
    project_data = data["Projects"]
    for proj in project_data.values():
        kloc = proj["SLOC"] / 1000.0
        effort = proj["Effort"] * 12
        time = proj["Development Time"] * 12
        kloc_values.append(kloc)
        effort_values.append(effort)
        time_values.append(time)
    return np.array(kloc_values), np.array(effort_values), np.array(time_values)

def fit_cocomo_effort(kloc_values, effort_values):
    """Fits COCOMO effort equation (Effort = a * KLOC^b) using log transformation."""
    scaler = MinMaxScaler()
    scaled_kloc = scaler.fit_transform(np.array(kloc_values).reshape(-1, 1))
    log_kloc = np.log(np.maximum(scaled_kloc.flatten(), 1e-10))  # Add a small constant to avoid log(0)
    log_effort = np.log(np.array(effort_values) + 1e-10)

    effort_model = Ridge(alpha=1.0)  # Add L2 regularization
    effort_model.fit(log_kloc.reshape(-1, 1), log_effort)

    b = effort_model.coef_[0]
    a = np.exp(effort_model.intercept_)

    # Ensure b is positive
    b = max(0.01, b)

    return a, b, effort_model.predict(log_kloc.reshape(-1, 1))

def fit_cocomo_time(effort_values, time_values):
    """Fits COCOMO time equation (Time = c * Effort^d) using log transformation."""
    log_effort = np.log(np.array(effort_values) + 1e-10)
    log_time = np.log(np.array(time_values) + 1e-10)

    time_model = Ridge(alpha=1.0)
    time_model.fit(log_effort.reshape(-1, 1), log_time)

    d = time_model.coef_[0]
    c = np.exp(time_model.intercept_)

    # Ensure d is positive
    d = max(0.01, d)

    return c, d, time_model.predict(log_effort.reshape(-1, 1))

def calculate_variance(actual_values, predicted_values):
    """Calculates the variance between actual and predicted values."""
    return np.var(actual_values - np.exp(predicted_values))

def tune_cocomo_parameters_cocomo_aware(kloc_values, effort_values, time_values, nominal_a, nominal_b, nominal_c, nominal_d, variation_range=0.1, cocomo_weight=0.1):
    """Tunes COCOMO parameters with COCOMO performance awareness."""
    best_a, best_b, best_c, best_d = nominal_a, nominal_b, nominal_c, nominal_d
    min_score = float('inf')

    # Wider range for a and b, narrower for c and d
    a_range = np.linspace(nominal_a * (1 - variation_range), nominal_a * (1 + variation_range * 2), 5)
    b_range = np.linspace(nominal_b * (1 - variation_range), nominal_b * (1 + variation_range * 2), 5)
    c_range = np.linspace(nominal_c * (1 - variation_range / 2), nominal_c * (1 + variation_range / 2), 5)
    d_range = np.linspace(nominal_d * (1 - variation_range / 2), nominal_d * (1 + variation_range / 2), 5)

    for a in a_range:
        for b in b_range:
            for c in c_range:
                for d in d_range:
                    _, _, predicted_log_effort = fit_cocomo_effort(kloc_values, effort_values)
                    effort_variance = calculate_variance(effort_values, predicted_log_effort)

                    _, _, predicted_log_time = fit_cocomo_time(effort_values, time_values)
                    time_variance = calculate_variance(time_values, predicted_log_time)

                    # Bias towards larger a and b, smaller c and d
                    cocomo_score = (nominal_a - a) + (nominal_b - b) + (d - nominal_d) + (c - nominal_c)

                    # Strongly penalize smaller a/b and larger c/d
                    score = effort_variance + time_variance + cocomo_weight * cocomo_score

                    if score < min_score:
                        min_score = score
                        best_a, best_b, best_c, best_d = a, b, c, d

    return best_a, best_b, best_c, best_d

def run_cocomo_analysis(data, cocomo_enum):
    """
    Runs COCOMO analysis using the COCOMO class and model type.
    """

    temp = COCOMO(0, cocomo_enum)

    kloc_values, effort_values, time_values = load_cocomo_data(data)
    nominal_a = temp.a #adjust these based on the cocomo model.
    nominal_b = temp.b
    nominal_c = temp.c
    nominal_d = temp.d

    tuned_a, tuned_b, tuned_c, tuned_d = tune_cocomo_parameters_cocomo_aware(
        kloc_values, effort_values, time_values, nominal_a, nominal_b, nominal_c, nominal_d
    )

    log_predicted_effort = np.log(tuned_a) + tuned_b * np.log(kloc_values + 1e-10)
    tuned_effort_prediction = np.exp(log_predicted_effort)

    log_predicted_time = np.log(tuned_c) + tuned_d * np.log(effort_values + 1e-10)
    tuned_time_prediction = np.exp(log_predicted_time)

    scaler = MinMaxScaler()
    scaled_effort_values = scaler.fit_transform(tuned_effort_prediction.reshape(-1, 1)).flatten()

    tuned_effort_variance = np.var(scaled_effort_values)
    tuned_time_variance = np.var(time_values)

    return tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_variance, tuned_time_variance, tuned_effort_prediction, tuned_time_prediction

def write_cocomo_to_excel(file_loc, tuned_a, tuned_b, tuned_c, tuned_d, tuned_effort_variance, tuned_time_variance, tuned_effort_prediction, tuned_time_prediction, data):
    """Writes COCOMO results to a new worksheet in the Excel file with the specified format, adds COCOMO(b) column, and removes unnecessary variances."""
    try:
        workbook = load_workbook(file_loc)
    except FileNotFoundError:
        workbook = Workbook()  # Create a new workbook if the file doesn't exist

    if "COCOMO Results" in workbook.sheetnames:
        del workbook["COCOMO Results"]  # Remove existing sheet if it exists
    worksheet = workbook.create_sheet("COCOMO Results")

    # Define styles
    bold_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font = Font(name='Arial', size=10)
    center_alignment = Alignment(horizontal='center')
    right_alignment = Alignment(horizontal='right')
    header_fill = PatternFill(start_color="B7C9E2", end_color="B7C9E2", fill_type="solid")  # Use B7C9E2 for light blue

    # Write COCOMO header (Starting at column B)
    start_col = 2  # Column B
    headers = ["Source Lines of Code", "Effort (Labor Months)", "COCOMO (a)", "COCOMO (b)", "KLOC^b", "Dev Time (Months) t_d", "COCOMO (c)", "Effort^d", "Predicted Effort", "Predicted Time"]
    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num + start_col - 1)
        cell.value = header
        cell.border = bold_border
        cell.alignment = center_alignment
        cell.fill = header_fill

    # Write COCOMO data (Starting at column B)
    kloc_values = [proj["SLOC"] / 1000.0 for proj in data["Projects"].values()]
    effort_values = [proj["Effort"] for proj in data["Projects"].values()]
    time_values = [proj["Development Time"] for proj in data["Projects"].values()]
    predicted_effort = tuned_a * (np.array(kloc_values) ** tuned_b)
    predicted_time = tuned_c * (np.array(effort_values) ** tuned_d)

    for row_num, (kloc, effort, time, pred_effort, pred_time) in enumerate(zip(kloc_values, effort_values, time_values, predicted_effort, predicted_time), start=2):
        worksheet.cell(row=row_num, column=start_col).value = kloc * 1000
        worksheet.cell(row=row_num, column=start_col + 1).value = effort
        worksheet.cell(row=row_num, column=start_col + 2).value = tuned_a
        worksheet.cell(row=row_num, column=start_col + 3).value = tuned_b
        worksheet.cell(row=row_num, column=start_col + 4).value = kloc ** tuned_b
        worksheet.cell(row=row_num, column=start_col + 5).value = time
        worksheet.cell(row=row_num, column=start_col + 6).value = tuned_c
        worksheet.cell(row=row_num, column=start_col + 7).value = effort ** tuned_d
        worksheet.cell(row=row_num, column=start_col + 8).value = pred_effort
        worksheet.cell(row=row_num, column=start_col + 9).value = pred_time
        for col_num in range(start_col, start_col + 10):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = bold_border
            cell.alignment = right_alignment

    # Calculate and write averages (Starting at column B)
    avg_row = row_num + 1
    worksheet.cell(row=avg_row, column=start_col - 1).value = "AVERAGE"
    worksheet.cell(row=avg_row, column=start_col - 1).border = bold_border
    worksheet.cell(row=avg_row, column=start_col - 1).fill = header_fill
    for col_num in range(start_col, start_col + 10):
        avg = np.mean([worksheet.cell(row=r, column=col_num).value for r in range(2, avg_row)])
        cell = worksheet.cell(row=avg_row, column=col_num)
        cell.value = avg
        cell.border = bold_border
        cell.alignment = right_alignment

    # Calculate and write variances (Starting at column B)
    var_row = row_num + 2
    worksheet.cell(row=var_row, column=start_col - 1).value = "Variance"
    worksheet.cell(row=var_row, column=start_col - 1).border = bold_border
    worksheet.cell(row=var_row, column=start_col - 1).fill = header_fill

    # Columns to skip variance calculation for
    skip_variance_cols = [start_col, start_col + 4, start_col + 8]

    for col_num in range(start_col, start_col + 10):
        if col_num in skip_variance_cols:
            worksheet.cell(row=var_row, column=col_num).value = ""  # Leave blank
        else:
            values = [worksheet.cell(row=r, column=col_num).value for r in range(2, avg_row)]
            variance = np.var(values)
            cell = worksheet.cell(row=var_row, column=col_num)
            cell.value = variance
            cell.border = bold_border
            cell.alignment = right_alignment

    # Auto-adjust column widths
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2  # Add some padding

    workbook.save(file_loc)

if __name__ == '__main__':

    if(len(argv) < 2):
        print("Missing Argument")
    else:
        main(argv[1:])
